import sys
import argparse
import ldap3
import openpyxl
import os
'''
    A simple python script for extracting user's displayname and email address
    from Active Directory.
    It uses ldap3 protocol to make connection to Domain controller, via port 389
    For training in Python, the script makes uses of class, staticmethod, class's attributes
    and also make use of argparse module for arguments passing from command line.
    Right now, it requires 5 arguments from command line, user,password, domain, port
    text file contains users and output excel file

    Usage:
        python <script file> 'domain\account' 'account password' 'FQDN' port_number 'textfile_contains_users ' 'excel_file_for_data'
        eg python .\usersinfo.py 'mydomain\robot' 'OOps1234567' 'mydomain.com.au' 389 '.\users.txt' '.\data.xlsx'
'''
# a simple function to display error when caught by th except block
def ErrorHandler(msg):
    print(msg)

'''
    creating server's connection via ldap3 protocol. Note that the function
    does not habe any hard data, everything are fed from parameters
'''
def ADBinding(user,passwd,domain,port):
    try:
        unswserver = ldap3.Server(domain,port=port,get_info='ALL')
        adconn = ldap3.Connection(unswserver,user=user,password=passwd,auto_bind=True,authentication='NTLM')
    except:
        ErrorHandler(sys.exc_info())
    else:
        return unswserver,adconn

'''
    The main class for operation. Note the class's attributes. There is no reason why
    we don't use instance's attributes, but for learning curve only.
    However, numInstances represent number of instances creation so it should be class's attr
'''
class ADOperation:
    numInstances=0
    adserver = None
    adConn = None
    zidsfile=None
    searchBase=None
    xldatafile=None
    def __init__(self):
        #whenever an instance's created, increment this counter
        ADOperation.numInstances +=1
    #declare the following method as static method. we don't need @staticmethod decorator for Py 3.7 and up
    @staticmethod
    def CheckNumberofInstances():
        if ADOperation.numInstances <=1:
            #these are arguments passing from command line
            parser = argparse.ArgumentParser(description='Getting user,passwd,domain and port from command line')
            parser.add_argument('user',help=r'enter your user account to connect server')
            parser.add_argument('passwd',help=r'enter password for your account')
            parser.add_argument('domain',help=r'enter domain you want to connect to via ldap')
            parser.add_argument('port',help=r'enter domain you want to connect to via ldap')
            parser.add_argument('zidsfile',help=r'Full Path to teh text file contain zids')
            parser.add_argument('xlfile',help=r'Full Path to teh text file contain zids')
            args = parser.parse_args()
            ADOperation.zidsfile = eval(args.zidsfile)
            ADOperation.searchBase = eval(args.domain)
            ADOperation.xldatafile = eval(args.xlfile)
            #callldap binding to the server
            ADOperation.adServer,ADOperation.adConn = ADBinding(eval(args.user),eval(args.passwd),eval(args.domain),eval(args.port))
    #GetUsersInfo method extract information from Active Directory
    def GetUsersInfo(self):
        try:
            #first check of the excel file we write to exist, if not create it
            if not os.path.exists(ADOperation.xldatafile):
                print('creating excel file now')
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet['A1']='1'
                wb.save(ADOperation.xldatafile)
                wb.close()
            '''
                domain passed in is in the form of 'mydomain.com.au', so we need to transform it
                to 'DC=mydomain,DC=com,DC=au'. ADOperation.searchBase.split('.'))) make a list
                ['mydomain','com','au'], we then map it to lambda function to tranform the result to
                'DC=mydomain,DC=com,DC=au'
            '''
            temp = list(map((lambda x: 'DC='+ x), ADOperation.searchBase.split('.')))
            searchbase = ','.join(temp)
            # ldap attributes we want to get from Active Directory. In this case we only
            #interested in displayName and email address
            ldapattr = ['displayName','title','department','company','proxyAddresses','lastLogon']
            #get workbook for writing
            workbook = openpyxl.load_workbook(ADOperation.xldatafile,read_only=False,data_only=True)
            #get the first worksheet
            sh = workbook.worksheets[0]
            currRow=0
            with open(ADOperation.zidsfile,'rt') as reader:
                print('Begin reading text file to execute each zid')
                for zid in reader:
                    currRow +=1
                    currCol =0
                    ldapfilterstr = "(&(objectclass=person)(samAccountName={0}))".format(zid)
                    ADOperation.adConn.search(searchbase,ldapfilterstr,attributes=ldapattr)
                    for attr in ldapattr:
                        currCol +=1
                        if attr == 'proxyAddresses':
                           email = ''.join([x for x in ADOperation.adConn.entries[0].proxyAddresses if 'SMTP:' in(str(x))])
                           sh.cell(row=currRow,column=currCol).value = email
                        else:
                            strtemp = 'ADOperation.adConn.entries[0].%s' % attr
                            sh.cell(row=currRow,column=currCol).value=str(eval(strtemp))
            workbook.save(ADOperation.xldatafile)
            workbook.close()
            print('Complete')
        except:
            ErrorHandler(sys.exc_info())

if __name__ == '__main__':
    unswad = ADOperation()
    ADOperation.CheckNumberofInstances()
    unswad.GetUsersInfo()




